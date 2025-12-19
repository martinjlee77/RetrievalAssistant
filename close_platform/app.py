import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import calendar
import numpy as np
import sys
import os
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from close_platform.db_config import get_connection, init_close_tables
from close_platform import qbo_connector

APP_TITLE = "Lynx Close Platform"

st.set_page_config(page_title=APP_TITLE,
                   page_icon="assets/images/favicon.ico",
                   layout="wide",
                   initial_sidebar_state="expanded")

try:
    st.logo("assets/images/nobkg.png", size="large")
except:
    pass

st.html("""
<style>
    .st-key-close_card {
        background: #36404A;
    }
    .st-key-close_card:hover {
        background: #46505A;
        cursor: pointer;
    }
</style>
""")


def get_all_months():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM close_monthly_close ORDER BY month_id DESC")
    rows = cursor.fetchall()
    conn.close()
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def create_new_month_close(month_str):
    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            "INSERT INTO close_monthly_close (month_id, status) VALUES (%s, 'Open')",
            (month_str, ))

        cursor.execute(
            "SELECT month_id FROM close_monthly_close WHERE month_id < %s ORDER BY month_id DESC LIMIT 1",
            (month_str, ))
        last_month_row = cursor.fetchone()

        if last_month_row:
            last_month_id = last_month_row['month_id']
            st.toast(f"Found history. Cloning tasks from {last_month_id}...",
                     icon="📋")
            cursor.execute(
                """
                INSERT INTO close_monthly_tasks (month_id, task_name, day_due, owner, instructions_link, task_note, status)
                SELECT %s, task_name, day_due, owner, instructions_link, task_note, 'Pending'
                FROM close_monthly_tasks
                WHERE month_id = %s
            """, (month_str, last_month_id))
        else:
            st.toast("First run detected. Using Master Template.", icon="🆕")
            cursor.execute(
                """
                INSERT INTO close_monthly_tasks (month_id, task_name, day_due, owner, task_note, status)
                SELECT %s, task_name, day_due, default_owner, task_note, 'Pending'
                FROM close_checklist_template
            """, (month_str, ))

        cursor.execute(
            """
            INSERT INTO close_monthly_balances (month_id, account_number, status)
            SELECT %s, account_number, 'Open'
            FROM close_accounts
        """, (month_str, ))

        conn.commit()
        st.toast(f"Successfully initialized {month_str}", icon="✅")

    except Exception as e:
        conn.rollback()
        if "duplicate key" in str(e).lower():
            st.error(f"Close for {month_str} already exists.")
        else:
            st.error(f"Error: {e}")
    finally:
        conn.close()


def update_task_status(task_id, new_status):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE close_monthly_tasks SET status = %s WHERE id = %s",
                   (new_status, task_id))
    conn.commit()
    conn.close()


def update_account_entry(entry_id, field, value):
    conn = get_connection()
    cursor = conn.cursor()
    query = f"UPDATE close_monthly_balances SET {field} = %s WHERE id = %s"
    cursor.execute(query, (value, entry_id))
    conn.commit()
    conn.close()


def update_permanent_link(account_number, link):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE close_accounts SET permanent_link = %s WHERE account_number = %s",
        (link, account_number))
    conn.commit()
    conn.close()


def get_last_sync_time(month_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT last_synced_at FROM close_monthly_close WHERE month_id = %s",
        (month_id, ))
    row = cursor.fetchone()
    conn.close()
    return row['last_synced_at'] if row else None


def get_account_group(acct_num):
    try:
        n = int(acct_num)
        if 10000 <= n <= 10999: return "01. Cash & Cash Equivalents"
        if 11000 <= n <= 11999: return "02. Accounts Receivable"
        if 12000 <= n <= 12999: return "03. Other Current Assets"
        if 13000 <= n <= 13999: return "04. Fixed Assets"
        if 14000 <= n <= 14999: return "03. Other Current Assets"
        if 20000 <= n <= 21999: return "05. Accounts Payable & Cards"
        if 22000 <= n <= 23999: return "06. Other Current Liabilities"
        if 26000 <= n <= 26999: return "06. Other Current Liabilities"
        if 24000 <= n <= 25999: return "07. Long-term Liabilities"
        if 27000 <= n <= 29999: return "07. Long-term Liabilities"
        if 30000 <= n <= 39999: return "08. Equity"
        if 40000 <= n <= 49999: return "09. Revenue"
        if 50000 <= n <= 59999: return "10. Cost of Sales"
        if 60000 <= n <= 69999: return "11. Operating Expenses"
        if 70000 <= n <= 89999: return "12. Other Income/Expense"
    except:
        pass
    return "13. Other / Unmapped"


def get_prior_month_id(curr_month_id):
    try:
        curr_date = datetime.strptime(curr_month_id, "%Y-%m")
        if curr_date.month == 1:
            prev_date = curr_date.replace(year=curr_date.year - 1, month=12)
        else:
            prev_date = curr_date.replace(month=curr_date.month - 1)
        return prev_date.strftime("%Y-%m")
    except:
        return None


def update_month_totals(month_id, debits, credits):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE close_monthly_close 
        SET qbo_total_debits = %s, qbo_total_credits = %s 
        WHERE month_id = %s
    """, (debits, credits, month_id))
    conn.commit()
    conn.close()


def get_month_totals(month_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT qbo_total_debits, qbo_total_credits FROM close_monthly_close WHERE month_id = %s",
        (month_id, ))
    row = cursor.fetchone()
    conn.close()
    return (row['qbo_total_debits'] or 0.0,
            row['qbo_total_credits'] or 0.0) if row else (0.0, 0.0)


def update_qbo_net_income(month_id, net_income):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE close_monthly_close 
        SET qbo_net_income = %s 
        WHERE month_id = %s
    """, (net_income, month_id))
    conn.commit()
    conn.close()


def get_qbo_net_income(month_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT qbo_net_income FROM close_monthly_close WHERE month_id = %s",
        (month_id, ))
    row = cursor.fetchone()
    conn.close()
    return row['qbo_net_income'] or 0.0 if row else 0.0


def get_business_day_delta(month_id):
    y, m = map(int, month_id.split('-'))
    last_day = calendar.monthrange(y, m)[1]
    month_end_date = date(y, m, last_day)
    today = date.today()
    if today <= month_end_date:
        return 0
    t_days = np.busday_count(month_end_date, today)
    return int(t_days)


def delete_month_close(month_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM close_monthly_tasks WHERE month_id = %s", (month_id,))
    cursor.execute("DELETE FROM close_monthly_balances WHERE month_id = %s", (month_id,))
    cursor.execute("DELETE FROM close_monthly_close WHERE month_id = %s", (month_id,))
    conn.commit()
    conn.close()


def render_lobby():
    st.title(f"📂 {APP_TITLE}")

    col_left, col_center, col_right = st.columns([1, 1, 1])
    with col_left:
        with st.expander("Start a New Month Close", expanded=False, width="stretch"):
            current_month = datetime.now().strftime("%Y-%m")
            new_month_input = st.text_input("Enter a New Month to Close", value=current_month)
            st.write("")
            if st.button("Initialize Close"):
                create_new_month_close(new_month_input)
                st.rerun()

    df_months = get_all_months()

    if df_months.empty:
        st.info("No close periods found. Initialize one above!")
        return

    st.subheader("Active Closes")

    cols = st.columns(4)
    for index, row in df_months.iterrows():
        col = cols[index % 4]
        with col:
            status_color = "⚪️" if row['status'] == 'Closed' else "🟢"
            lock_icon = "🔒 Closed" if row['is_locked'] else "🟢 Unlocked"
            is_open = row['status'] != 'Closed'

            with st.container(border=True,
                              key=f"close_card_{row['month_id']}"):
                st.markdown(f"### {row['month_id']}")
                st.markdown(
                    f"**Close Status:** {status_color} {row['status']}")
                st.markdown(f"**QBO:** {lock_icon}")

                btn_col1, btn_col2 = st.columns(2)
                with btn_col1:
                    if st.button(f":orange[Open]",
                                 key=f"btn_{row['month_id']}",
                                 use_container_width=True):
                        st.session_state['active_month'] = row['month_id']
                        st.rerun()
                with btn_col2:
                    if is_open:
                        with st.popover("🗑️", use_container_width=True):
                            st.info(f"Delete {row['month_id']}? This cannot be undone.")
                            if st.button("Yes, Delete",
                                         key=f"confirm_del_{row['month_id']}",
                                         type="primary"):
                                delete_month_close(row['month_id'])
                                st.rerun()


def render_checklist_tab(month_id, owner_filter):
    conn = get_connection()
    cursor = conn.cursor()

    query = "SELECT * FROM close_monthly_tasks WHERE month_id = %s"
    params = [month_id]

    if owner_filter != "All":
        query += " AND owner = %s"
        params.append(owner_filter)

    query += " ORDER BY day_due ASC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    df = pd.DataFrame(rows) if rows else pd.DataFrame()

    if df.empty:
        st.info(
            "No tasks found for this month. Add tasks to the template or clone from a prior month."
        )
        return

    current_t_day = get_business_day_delta(month_id)
    total_tasks = len(df)
    done_tasks = len(df[df['status'].isin(['Done', 'N/A'])])
    prog = done_tasks / total_tasks if total_tasks > 0 else 0

    c1, c2, c3 = st.columns([5, 1, 2])

    with c1:
        st.markdown(f"### 📅 Today is T+{current_t_day}")
        st.caption("✅ Done | 🔴 Late | 🟡 Due Today | ⚪️ Upcoming")

    with c2:
        st.metric("Completion", f"{done_tasks}/{total_tasks}")

    with c3:
        save_clicked = st.button("**Please Save Manually**",
                                 icon="💾",
                                 type="primary",
                                 use_container_width=True,
                                 help="Save Changes")

    st.progress(prog, text=f"Overall Completion: {int(prog*100)}%")

    def get_alert_icon(row):
        if row['status'] in ['Done', 'N/A']:
            return "✅"
        if row['day_due'] < current_t_day:
            return "🔴"
        elif row['day_due'] == current_t_day:
            return "🟡"
        else:
            return "⚪️"

    df['Health'] = df.apply(get_alert_icon, axis=1)

    st.divider()

    ROLES = ["VP Accounting", "Ops", "CFO", "Accounting Firm"]
    STATUS_OPTS = ["Pending", "In Progress", "Done", "N/A"]

    column_config = {
        "id":
        None,
        "month_id":
        None,
        "Health":
        st.column_config.TextColumn("", width=10, disabled=True),
        "day_due":
        st.column_config.NumberColumn("Due",
                                      width=25,
                                      min_value=1,
                                      max_value=20,
                                      format="%d"),
        "task_name":
        st.column_config.TextColumn("Task Name", width="large", required=True),
        "owner":
        st.column_config.SelectboxColumn("Owner", options=ROLES,
                                         required=True),
        "instructions_link":
        st.column_config.LinkColumn("SOP", display_text="Open"),
        "task_note":
        st.column_config.TextColumn("Notes"),
        "status":
        st.column_config.SelectboxColumn("Status",
                                         options=STATUS_OPTS,
                                         required=True)
    }

    display_df = df[[
        'id', 'Health', 'day_due', 'task_name', 'owner',
        'instructions_link', 'task_note', 'status', 'month_id'
    ]]

    dynamic_height = min((len(df) + 1) * 35, 1500)

    edited_df = st.data_editor(display_df,
                               key="checklist_editor",
                               hide_index=True,
                               column_config=column_config,
                               use_container_width=True,
                               num_rows="dynamic",
                               height=dynamic_height)

    if save_clicked:
        conn = get_connection()
        cursor = conn.cursor()

        original_ids = set(df['id'].dropna().astype(int))
        current_ids = set(edited_df['id'].dropna().astype(int))
        deleted_ids = original_ids - current_ids

        for del_id in deleted_ids:
            cursor.execute("DELETE FROM close_monthly_tasks WHERE id = %s",
                           (del_id, ))

        for index, row in edited_df.iterrows():
            if pd.notna(row['id']):
                cursor.execute(
                    """
                    UPDATE close_monthly_tasks 
                    SET task_name=%s, day_due=%s, owner=%s, instructions_link=%s, task_note=%s, status=%s
                    WHERE id=%s
                """, (row['task_name'], row['day_due'],
                      row['owner'], row['instructions_link'], row['task_note'], row['status'],
                      row['id']))
            else:
                cursor.execute(
                    """
                    INSERT INTO close_monthly_tasks (month_id, task_name, day_due, owner, instructions_link, task_note, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (month_id, row['task_name'], row['day_due'],
                      row['owner'], row['instructions_link'], row['task_note'], row['status']))

        conn.commit()
        conn.close()
        st.toast("Checklist Updated!", icon="✅")
        st.rerun()


def render_tb_tab(month_id):
    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT 
            mb.id, 
            mb.account_number, 
            a.account_name, 
            a.permanent_link,
            mb.qbo_balance, 
            mb.expected_balance, 
            mb.status,
            mb.rec_note
        FROM close_monthly_balances mb
        JOIN close_accounts a ON mb.account_number = a.account_number
        WHERE mb.month_id = %s
        ORDER BY mb.account_number ASC
    """
    cursor.execute(query, (month_id, ))
    rows = cursor.fetchall()
    conn.close()

    df_bal = pd.DataFrame(rows) if rows else pd.DataFrame()

    if df_bal.empty:
        st.warning(
            "No account data found. Connect to QBO and sync data, or add accounts manually."
        )
        return

    df_bal['Group'] = df_bal['account_number'].apply(get_account_group)

    total_accts = len(df_bal)
    done_accts = len(df_bal[df_bal['status'].isin(['Supported', 'Reviewed'])])
    progress = done_accts / total_accts if total_accts > 0 else 0

    st.progress(
        progress,
        text=
        f"Substantiation Progress: {int(progress*100)}% ({done_accts}/{total_accts})"
    )
    st.divider()

    groups = sorted(df_bal['Group'].unique())

    for group in groups:
        with st.expander(f"📂 {group}", expanded=False):
            group_rows = df_bal[df_bal['Group'] == group].copy()

            group_rows['Variance'] = group_rows['qbo_balance'] - group_rows[
                'expected_balance']
            group_rows['qbo_balance_fmt'] = group_rows['qbo_balance'].apply(
                lambda x: "{:,.2f}".format(x))
            group_rows['expected_balance_fmt'] = group_rows['expected_balance'].apply(
                lambda x: "{:,.2f}".format(x))
            group_rows['Variance_fmt'] = group_rows['Variance'].apply(
                lambda x: "{:,.2f}".format(x))

            column_config = {
                "id":
                None,
                "Group":
                None,
                "qbo_balance":
                None,
                "expected_balance_fmt":
                None,
                "Variance":
                None,
                "account_number":
                "Acct #",
                "account_name":
                "Account Name",
                "permanent_link":
                st.column_config.LinkColumn("Workpaper Link",
                                            display_text="Open Folder"),
                "rec_note":
                st.column_config.TextColumn("Notes",
                                            width="medium"),
                "qbo_balance_fmt":
                st.column_config.TextColumn("QBO Bal", disabled=True),
                "expected_balance":
                st.column_config.NumberColumn("Exp Bal", format="%.2f"),
                "Variance_fmt":
                st.column_config.TextColumn("Diff", disabled=True),
                "status":
                st.column_config.SelectboxColumn(
                    "Status",
                    options=["Open", "Supported", "Reviewed"],
                    required=True)
            }

            display_cols = [
                'id', 'account_number', 'account_name', 'qbo_balance_fmt',
                'expected_balance', 'Variance_fmt', 'status', 'permanent_link',
                'rec_note'
            ]

            edited_df = st.data_editor(group_rows[display_cols],
                                       key=f"editor_{group}",
                                       hide_index=True,
                                       column_config=column_config,
                                       use_container_width=True,
                                       height=min((len(group_rows) + 1) * 35,
                                                  600))

            changes_detected = False
            for index, row in edited_df.iterrows():
                orig_row = df_bal[df_bal['id'] == row['id']].iloc[0]

                if abs(row['expected_balance'] -
                       orig_row['expected_balance']) > 0.001:
                    update_account_entry(row['id'], 'expected_balance',
                                         row['expected_balance'])
                    changes_detected = True

                if row['rec_note'] != orig_row['rec_note']:
                    conn = get_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE close_monthly_balances SET rec_note = %s WHERE id = %s",
                        (row['rec_note'], row['id']))
                    conn.commit()
                    conn.close()
                    changes_detected = True

                if row['permanent_link'] != orig_row['permanent_link']:
                    update_permanent_link(row['account_number'],
                                          row['permanent_link'])
                    changes_detected = True

                if row['status'] != orig_row['status']:
                    new_variance = orig_row['qbo_balance'] - row[
                        'expected_balance']

                    if row['status'] in ['Supported', 'Reviewed'
                                         ] and abs(new_variance) > 0.01:
                        st.toast(
                            f"❌ Account {row['account_number']}: Cannot mark complete. Variance is {new_variance:,.2f}",
                            icon="🚫")

                        if f"editor_{group}" in st.session_state:
                            del st.session_state[f"editor_{group}"]

                        st.rerun()
                    else:
                        update_account_entry(row['id'], 'status',
                                             row['status'])
                        changes_detected = True

            if changes_detected:
                st.rerun()


def render_flux_tab(month_id, threshold_amt, threshold_pct):
    conn = get_connection()
    cursor = conn.cursor()
    prev_month_id = get_prior_month_id(month_id)

    query = """
        SELECT 
            curr.id,
            curr.account_number,
            a.account_name,
            curr.qbo_balance as curr_bal,
            prev.qbo_balance as prev_bal,
            curr.variance_note
        FROM close_monthly_balances curr
        LEFT JOIN close_monthly_balances prev 
            ON curr.account_number = prev.account_number 
            AND prev.month_id = %s
        JOIN close_accounts a ON curr.account_number = a.account_number
        WHERE curr.month_id = %s
        ORDER BY curr.account_number ASC
    """
    cursor.execute(query, (prev_month_id, month_id))
    rows = cursor.fetchall()

    cursor.execute(
        "SELECT SUM(qbo_balance) as total FROM close_monthly_balances WHERE month_id = %s",
        (month_id, ))
    sub_ledger_result = cursor.fetchone()
    sub_ledger_total = sub_ledger_result['total'] if sub_ledger_result and sub_ledger_result.get('total') else 0.0
    
    conn.close()
    
    y, m = int(month_id[:4]), int(month_id[5:7])
    last_day = calendar.monthrange(y, m)[1]
    month_end_date_str = f"{month_id}-{last_day:02d}"
    ytd_net_income_from_qbo = qbo_connector.get_ytd_net_income(month_end_date_str)

    df = pd.DataFrame(rows) if rows else pd.DataFrame()

    if df.empty:
        st.info("No data found.")
        return

    df['Group'] = df['account_number'].apply(get_account_group)
    df['prev_bal'] = df['prev_bal'].fillna(0.0)
    df['diff_amt'] = df['curr_bal'] - df['prev_bal']
    df['diff_pct'] = df.apply(lambda x: (x['diff_amt'] / x['prev_bal'] * 100)
                              if x['prev_bal'] != 0 else 0.0,
                              axis=1)

    count_db = len(df)
    count_tab2 = len(df[df['Group'] != "13. Other / Unmapped"])
    count_tab3 = len(df)
    is_pop_match = (count_db == count_tab2 == count_tab3)

    total_assets = df[df['Group'].str.startswith(
        ('01', '02', '03', '04'))]['curr_bal'].sum()
    total_liabs = df[df['Group'].str.startswith(
        ('05', '06', '07'))]['curr_bal'].sum()
    total_equity = df[df['Group'].str.startswith(('08'))]['curr_bal'].sum()
    ytd_net_income = ytd_net_income_from_qbo
    
    bs_check_val = total_assets + total_liabs + total_equity + ytd_net_income
    is_bs_balanced = abs(bs_check_val) < 1000

    df_mapped = df[df['Group'] != "13. Other / Unmapped"]
    
    df['acct_num_int'] = df['account_number'].apply(lambda x: int(x) if str(x).isdigit() else 0)
    df_pnl_all = df[df['acct_num_int'] >= 40000]
    df_pnl_mapped = df_pnl_all[df_pnl_all['Group'] != "13. Other / Unmapped"]
    
    tab2_month_ni = df_pnl_mapped['curr_bal'].sum()
    tab3_month_ni = df_pnl_all['curr_bal'].sum()
    saved_qbo_ni = get_qbo_net_income(month_id)

    internal_ni_match = abs(tab2_month_ni - tab3_month_ni) < 1.0
    external_ni_match = abs(tab3_month_ni - saved_qbo_ni) < 1000
    is_tied_out = internal_ni_match and saved_qbo_ni != 0.0 and external_ni_match

    st.markdown("### 🛡 Audit Control Center")

    r1_c1, r1_c2 = st.columns(2)

    with r1_c1:
        with st.container(border=True):
            st.markdown("**1. Population Integrity Check**")
            st.caption("Verifies DB rows match visible accounts in Tab 2 & 3.")

            c1, c2, c3 = st.columns(3)
            with c1:
                st.caption("Raw DB")
                st.markdown(f"**{count_db}**")
            with c2:
                st.caption("Tab 2")
                st.markdown(f"**{count_tab2}**")
            with c3:
                st.caption("Tab 3")
                st.markdown(f"**{count_tab3}**")

            st.write("")
            if is_pop_match:
                st.success("✅ Valid")
            else:
                st.error(f"🚨 Variance: {count_db - count_tab2}")

    with r1_c2:
        with st.container(border=True):
            st.markdown("**2. Balance Sheet Logic Check**")
            st.caption("Assets + Liabilities + Equity + YTD NI = 0")

            c1, c2, c3, c4, c5 = st.columns(5)
            with c1:
                st.caption("Assets")
                st.markdown(f"**{total_assets:,.0f}**")
            with c2:
                st.caption("Liabilities")
                st.markdown(f"**{total_liabs:,.0f}**")
            with c3:
                st.caption("Equity")
                st.markdown(f"**{total_equity:,.0f}**")
            with c4:
                st.caption("YTD NI")
                st.markdown(f"**{ytd_net_income:,.0f}**")
            with c5:
                st.caption("Net Check")
                st.markdown(f"**{bs_check_val:,.2f}**")

            st.write("")
            if is_bs_balanced:
                st.success("✅ Balanced")
            else:
                st.error("🚨 Out of Balance")

    with st.container(border=True):
        st.markdown("**3. Income Statement Tie-Out**")
        st.caption(
            "Current month Net Income: Tab 2 vs. Tab 3 vs. QBO P&L Report")

        c_lbl, c_t2, c_t3, c_qbo = st.columns([1.2, 1.2, 1.2, 2])

        c_lbl.markdown("###### ")
        c_t2.markdown("###### Tab 2")
        c_t3.markdown("###### Tab 3")
        c_qbo.markdown("###### QBO Input")

        c_lbl.markdown("Net Income")
        c_t2.markdown(f"**{tab2_month_ni:,.2f}**")
        c_t3.markdown(f"**{tab3_month_ni:,.2f}**")
        val_ni = c_qbo.number_input("Net Income from QBO P&L",
                                    value=saved_qbo_ni,
                                    key="aud_ni",
                                    label_visibility="collapsed",
                                    format="%.2f")

        if round(val_ni, 2) != round(saved_qbo_ni, 2):
            update_qbo_net_income(month_id, val_ni)
            st.rerun()

        if is_tied_out:
            st.success("✅ Income Statement Matched")
        else:
            if not internal_ni_match:
                st.error(f"🚨 Internal Variance: {tab2_month_ni - tab3_month_ni:,.2f}")

            if saved_qbo_ni != 0.0 and not external_ni_match:
                st.error(f"🚨 External Variance: {tab3_month_ni - saved_qbo_ni:,.2f}")

    st.divider()

    st.subheader("📊 Flux Analysis")

    updates_made = False
    for index, row in df.iterrows():
        is_material = (abs(row['diff_amt']) >= threshold_amt) and (abs(
            row['diff_pct']) >= threshold_pct)
        current_note = str(
            row['variance_note']) if row['variance_note'] else ""
        if not is_material and current_note == "":
            update_account_entry(row['id'], 'variance_note', "N/A")
            df.at[index, 'variance_note'] = "N/A"
            updates_made = True
    if updates_made:
        st.rerun()

    display_df = df.copy()
    display_df['prev_bal_fmt'] = display_df['prev_bal'].apply(
        lambda x: "{:,.2f}".format(x))
    display_df['curr_bal_fmt'] = display_df['curr_bal'].apply(
        lambda x: "{:,.2f}".format(x))
    display_df['diff_amt_fmt'] = display_df['diff_amt'].apply(
        lambda x: "{:,.2f}".format(x))
    display_df['diff_pct_fmt'] = display_df['diff_pct'].apply(
        lambda x: "{:.1f}%".format(x))

    def get_status_icon(row):
        orig_row = df[df['id'] == row['id']].iloc[0]
        is_material = (abs(orig_row['diff_amt']) >= threshold_amt) and (abs(
            orig_row['diff_pct']) >= threshold_pct)
        note = str(orig_row['variance_note']) if orig_row['variance_note'] else ""
        if not is_material:
            return "⚪️"
        elif is_material and (note == "" or note == "None"):
            return "🔴"
        else:
            return "✅"

    display_df['Action'] = display_df.apply(get_status_icon, axis=1)

    column_config = {
        "id": None,
        "Group": None,
        "prev_bal": None,
        "curr_bal": None,
        "diff_amt": None,
        "diff_pct": None,
        "account_number": "Acct #",
        "account_name": "Account Name",
        "prev_bal_fmt": st.column_config.TextColumn("Prior Month", disabled=True),
        "curr_bal_fmt": st.column_config.TextColumn("Current Month", disabled=True),
        "diff_amt_fmt": st.column_config.TextColumn("Var $", disabled=True),
        "diff_pct_fmt": st.column_config.TextColumn("Var %", disabled=True),
        "variance_note": st.column_config.TextColumn("Explanation",
                                                     width="large"),
        "Action": st.column_config.TextColumn("St",
                                              width="small",
                                              disabled=True)
    }

    final_view = display_df[[
        'id', 'account_number', 'account_name', 'prev_bal_fmt', 'curr_bal_fmt',
        'diff_amt_fmt', 'diff_pct_fmt', 'Action', 'variance_note'
    ]]

    edited_df = st.data_editor(final_view,
                               key=f"audit_flux_editor_{month_id}",
                               hide_index=True,
                               column_config=column_config,
                               use_container_width=True,
                               height=min((len(df) + 1) * 35, 1200))

    changes_detected = False
    for index, row in edited_df.iterrows():
        original_note = str(df.at[index, 'variance_note'])
        new_note = str(row['variance_note'])
        if original_note != new_note:
            update_account_entry(row['id'], 'variance_note', new_note)
            changes_detected = True

    if changes_detected:
        st.toast("Saved!", icon="💾")
        st.rerun()


def generate_excel_export(month_id):
    from openpyxl import Workbook

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT task_name, day_due, owner, instructions_link, task_note, status
        FROM close_monthly_tasks WHERE month_id = %s ORDER BY day_due ASC
    """, (month_id, ))
    tasks_data = cursor.fetchall()
    df_tasks = pd.DataFrame(tasks_data,
                            columns=[
                                'Task Name', 'Due (T+)', 'Owner',
                                'SOP Link', 'Notes', 'Status'
                            ]) if tasks_data else pd.DataFrame()

    cursor.execute(
        """
        SELECT 
            mb.account_number,
            a.account_name,
            mb.qbo_balance,
            mb.expected_balance,
            (mb.qbo_balance - mb.expected_balance) as variance,
            mb.status,
            mb.rec_note,
            a.permanent_link
        FROM close_monthly_balances mb
        JOIN close_accounts a ON mb.account_number = a.account_number
        WHERE mb.month_id = %s
        ORDER BY mb.account_number ASC
    """, (month_id, ))
    tb_data = cursor.fetchall()
    df_tb = pd.DataFrame(
        tb_data,
        columns=[
            'Acct #', 'Account Name', 'QBO Balance', 'Expected Balance',
            'Variance', 'Status', 'Rec Note', 'Workpaper Link'
        ]) if tb_data else pd.DataFrame()

    prev_month_id = get_prior_month_id(month_id)
    cursor.execute(
        """
        SELECT 
            curr.account_number,
            a.account_name,
            prev.qbo_balance as prev_bal,
            curr.qbo_balance as curr_bal,
            (curr.qbo_balance - COALESCE(prev.qbo_balance, 0)) as diff_amt,
            curr.variance_note
        FROM close_monthly_balances curr
        LEFT JOIN close_monthly_balances prev 
            ON curr.account_number = prev.account_number 
            AND prev.month_id = %s
        JOIN close_accounts a ON curr.account_number = a.account_number
        WHERE curr.month_id = %s
        ORDER BY curr.account_number ASC
    """, (prev_month_id, month_id))
    flux_data = cursor.fetchall()
    conn.close()

    df_flux = pd.DataFrame(flux_data,
                           columns=[
                               'Acct #', 'Account Name', 'Prior Month',
                               'Current Month', 'Variance $', 'Explanation'
                           ]) if flux_data else pd.DataFrame()
    if not df_flux.empty:
        df_flux['Prior Month'] = df_flux['Prior Month'].fillna(0)
        df_flux['Variance %'] = df_flux.apply(
            lambda x: (x['Variance $'] / x['Prior Month'] * 100)
            if x['Prior Month'] != 0 else 0.0,
            axis=1)
        df_flux = df_flux[[
            'Acct #', 'Account Name', 'Prior Month', 'Current Month',
            'Variance $', 'Variance %', 'Explanation'
        ]]

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="36404A",
                              end_color="36404A",
                              fill_type="solid")
    header_alignment = Alignment(horizontal="center",
                                 vertical="center",
                                 wrap_text=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws_tasks = wb.active
    ws_tasks.title = "Checklist"
    if not df_tasks.empty:
        for r_idx, row in enumerate(
                dataframe_to_rows(df_tasks, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_tasks.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        for col in ws_tasks.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_tasks.column_dimensions[col[0].column_letter].width = min(
                max_length + 2, 50)

    ws_tb = wb.create_sheet(title="Trial Balance")
    if not df_tb.empty:
        for r_idx, row in enumerate(
                dataframe_to_rows(df_tb, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_tb.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        for col in ws_tb.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_tb.column_dimensions[col[0].column_letter].width = min(
                max_length + 2, 50)

    ws_flux = wb.create_sheet(title="Flux Analysis")
    if not df_flux.empty:
        for r_idx, row in enumerate(
                dataframe_to_rows(df_flux, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_flux.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        for col in ws_flux.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_flux.column_dimensions[col[0].column_letter].width = min(
                max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def render_workspace(month_id):
    with st.sidebar:
        st.header(f"🗓 {month_id}")
        if st.button("← Back to Lobby"):
            del st.session_state['active_month']
            st.rerun()

        st.divider()
        st.subheader("Filters")
        owner_filter = st.selectbox(
            "Task Owner",
            ["All", "VP Accounting", "Ops", "CFO", "Accounting Firm"])

        st.divider()
        st.subheader("Flux Thresholds")
        thresh_amt = st.number_input("Min Variance ($)", value=5000, step=1000)
        thresh_pct = st.number_input("Min Variance (%)", value=10, step=1)

        st.divider()
        st.subheader("Export")
        excel_data = generate_excel_export(month_id)
        st.download_button(
            label="📥 Download Close Package",
            data=excel_data,
            file_name=f"close_package_{month_id}.xlsx",
            mime=
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

        st.divider()
        st.subheader("Data Sync")

        has_encryption_key = qbo_connector._has_encryption_key()

        if not has_encryption_key:
            st.warning(
                "⚠️ QBO_ENCRYPTION_KEY not set. Add this secret before connecting to QuickBooks."
            )
            client = None
        else:
            try:
                client = qbo_connector.get_active_client()
            except Exception:
                client = None

        if client:
            st.success("✅ QBO Connected")

            last_sync = get_last_sync_time(month_id)
            if last_sync:
                st.caption(f"Last Sync: {last_sync}")
            else:
                st.caption("Last Sync: Never")

            if st.button("🔄 Sync Live Data"):
                with st.spinner("Fetching Trial Balance from QBO..."):
                    try:
                        y, m = map(int, month_id.split('-'))
                        last_day = calendar.monthrange(y, m)[1]
                        end_date_str = f"{month_id}-{last_day}"

                        qbo_data = qbo_connector.fetch_trial_balance(
                            end_date_str)

                        if qbo_data:
                            conn = get_connection()
                            cursor = conn.cursor()

                            new_accounts_found = 0
                            updated_balances = 0

                            for acct_num, details in qbo_data.items():
                                name = details['name']
                                bal = details['balance']

                                cursor.execute(
                                    "SELECT account_number FROM close_accounts WHERE account_number = %s",
                                    (acct_num, ))
                                if not cursor.fetchone():
                                    cat = 'BS' if int(
                                        acct_num) < 40000 else 'PL'
                                    cursor.execute(
                                        "INSERT INTO close_accounts (account_number, account_name, category, permanent_link) VALUES (%s, %s, %s, '')",
                                        (acct_num, name, cat))
                                    new_accounts_found += 1

                                cursor.execute(
                                    "SELECT id FROM close_monthly_balances WHERE month_id = %s AND account_number = %s",
                                    (month_id, acct_num))
                                row = cursor.fetchone()

                                if row:
                                    cursor.execute(
                                        "UPDATE close_monthly_balances SET qbo_balance = %s WHERE id = %s",
                                        (bal, row['id']))
                                else:
                                    cursor.execute(
                                        "INSERT INTO close_monthly_balances (month_id, account_number, qbo_balance, status) VALUES (%s, %s, %s, 'Open')",
                                        (month_id, acct_num, bal))
                                updated_balances += 1

                            now_str = datetime.now().strftime(
                                "%Y-%m-%d %H:%M:%S")
                            cursor.execute(
                                "UPDATE close_monthly_close SET last_synced_at = %s WHERE month_id = %s",
                                (now_str, month_id))

                            conn.commit()
                            conn.close()

                            if new_accounts_found > 0:
                                st.toast(
                                    f"Auto-discovered {new_accounts_found} new accounts!"
                                )
                            st.toast(f"Synced {updated_balances} balances.",
                                     icon="✅")
                            st.rerun()
                        else:
                            st.error("Failed to fetch data.")
                    except Exception as e:
                        st.error(f"Sync Error: {e}")
        elif has_encryption_key:
            st.warning("Not Connected")
            
            with st.expander("📋 Manual Token Entry (from OAuth Playground)"):
                st.caption("Get tokens from Intuit OAuth 2.0 Playground, then paste them here:")
                manual_realm = st.text_input("Realm ID", key="manual_realm")
                manual_access = st.text_area("Access Token", key="manual_access", height=100)
                manual_refresh = st.text_input("Refresh Token", key="manual_refresh")
                
                if st.button("💾 Save Tokens", type="primary"):
                    if manual_realm and manual_access and manual_refresh:
                        try:
                            qbo_connector.save_manual_tokens(manual_realm, manual_access, manual_refresh)
                            st.success("Tokens saved! Refreshing...")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Failed to save tokens: {e}")
                    else:
                        st.warning("Please fill in all three fields.")
            
            auth_url = qbo_connector.get_auth_url()
            if auth_url:
                with st.expander("🔗 OAuth Login (Alternative)"):
                    st.link_button("Login to Intuit", auth_url)
                    st.caption("Note: Requires redirect URI configured in Intuit Developer Portal")

    st.title(f"Close Workspace: {month_id}")

    tab1, tab2, tab3 = st.tabs([
        "📋 Process (Checklist)", "⚖️ Substantiation (TB)", "📊 Reporting (Flux)"
    ])

    with tab1:
        render_checklist_tab(month_id, owner_filter)

    with tab2:
        render_tb_tab(month_id)

    with tab3:
        render_flux_tab(month_id, thresh_amt, thresh_pct)


def main():
    try:
        init_close_tables()
    except Exception as e:
        st.error(f"Database initialization error: {e}")
        return

    query_params = st.query_params
    if "code" in query_params:
        auth_code = query_params["code"]
        realm_id = query_params["realmId"]
        try:
            qbo_connector.handle_callback(auth_code, realm_id)
            st.toast("Successfully Connected to QuickBooks!", icon="🔗")
            st.query_params.clear()
        except Exception as e:
            st.error(f"Login Failed: {e}")

    if 'active_month' in st.session_state:
        render_workspace(st.session_state['active_month'])
    else:
        render_lobby()


if __name__ == '__main__':
    main()
